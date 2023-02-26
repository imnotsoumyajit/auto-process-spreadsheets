import openpyxl as xl #for xlsx files 

from openpyxl.chart import BarChart,Reference #for chart 


wb=xl.load_workbook('transactions.xlsx') #this is static and will only work for this file so..we do this in a function
# def process_workbook(filename):


sheet = wb['Sheet1']
# cell = sheet['a1']
# cell=sheet.cell(1,1)
# print(cell.value)
# print(sheet.max_row)

for row in range(2 , sheet.max_row+1):
    # We use +1 as range will not take the last value otherwise 
    # print(row)
    cell=sheet.cell(row,3)
    # print(cell.value)
    corrected_price=cell.value*0.9
    # add all the corrected prices in a new column(4th column)
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price

#to select a range of values we use 'Reference'
values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4,
          )
# Adding chart 
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

# Save in another file , dont override the original file 
wb.save('transactions2.xlsx')

